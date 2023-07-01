# pip install beautifulsoup4 : 스크래핑을 위한 패키지
# pip install lxml : 구문 분석 파서하는 패키지
# pip install openpyxl : 엑셀 사용하는 패키지
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup

urls=["https://www.saramin.co.kr/zf_user/search?loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&cat_kewd=277%2C236&exp_cd=1&exp_none=y&edu_min=6&edu_max=10&edu_none=y&panel_type=&search_optional_item=y&search_done=y&panel_count=y&preview=y&searchword=&show_applied=&except_read=&ai_head_hunting=&mainSearch=n&recruitPage=1&recruitSort=relation&recruitPageCount=100&inner_com_type=&company_cd=0%2C1%2C2%2C3%2C4%2C5%2C6%2C7%2C9%2C10&quick_apply=","https://www.saramin.co.kr/zf_user/search?loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&cat_kewd=277%2C236&exp_cd=1&exp_none=y&edu_min=6&edu_max=10&edu_none=y&panel_type=&search_optional_item=y&search_done=y&panel_count=y&preview=y&searchword=&show_applied=&except_read=&ai_head_hunting=&mainSearch=n&recruitPage=2&recruitSort=relation&recruitPageCount=100&inner_com_type=&company_cd=0%2C1%2C2%2C3%2C4%2C5%2C6%2C7%2C9%2C10&quick_apply=","https://www.saramin.co.kr/zf_user/search?loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&cat_kewd=277%2C236&exp_cd=1&exp_none=y&edu_min=6&edu_max=10&edu_none=y&panel_type=&search_optional_item=y&search_done=y&panel_count=y&preview=y&searchword=&show_applied=&except_read=&ai_head_hunting=&mainSearch=n&recruitPage=3&recruitSort=relation&recruitPageCount=100&inner_com_type=&company_cd=0%2C1%2C2%2C3%2C4%2C5%2C6%2C7%2C9%2C10&quick_apply=","https://www.saramin.co.kr/zf_user/search?loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&cat_kewd=277%2C236&exp_cd=1&exp_none=y&edu_min=6&edu_max=10&edu_none=y&panel_type=&search_optional_item=y&search_done=y&panel_count=y&preview=y&searchword=&show_applied=&except_read=&ai_head_hunting=&mainSearch=n&recruitPage=4&recruitSort=relation&recruitPageCount=100&inner_com_type=&company_cd=0%2C1%2C2%2C3%2C4%2C5%2C6%2C7%2C9%2C10&quick_apply=","https://www.saramin.co.kr/zf_user/search?loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&cat_kewd=277%2C236&exp_cd=1&exp_none=y&edu_min=6&edu_max=10&edu_none=y&panel_type=&search_optional_item=y&search_done=y&panel_count=y&preview=y&searchword=&show_applied=&except_read=&ai_head_hunting=&mainSearch=n&recruitPage=5&recruitSort=relation&recruitPageCount=100&inner_com_type=&company_cd=0%2C1%2C2%2C3%2C4%2C5%2C6%2C7%2C9%2C10&quick_apply="]
column_index = 2
for url in urls:
    headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
    res = requests.get(url, headers=headers)
    res.raise_for_status()

    soup = BeautifulSoup(res.text, "lxml")

    companys = soup.find_all("strong", attrs={"class":"corp_name"})

    row_index = 2
    wb = load_workbook("company.xlsx")
    ws = wb.active
    # ws.title = "기업"
    for company in companys:
        c = company.get_text().strip()
        ws.cell(row=row_index, column=column_index, value=c)
        row_index = row_index + 1
    column_index = column_index + 2
    wb.save("company.xlsx")