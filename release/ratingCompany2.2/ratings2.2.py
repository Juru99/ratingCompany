import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup

wb = load_workbook('./result/company2.2(2023-07-29).xlsx')
ws = wb.active
row_index = 2
keywords = []
while not ws.cell(row = row_index, column = 1).value is None:
  keywords.append(ws.cell(row = row_index, column = 1).value)
  row_index = row_index + 1
row_index = 2

headers = {
    "User-Agent": "Mozilla/5.0 (X11; CrOS x86_64 12871.102.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.141 Safari/537.36",
    "Connection": "close"}

for keyword in keywords:
  res = requests.get(f'https://www.jobplanet.co.kr/search?query={keyword}&category=search_new&search_keyword_hint_id=&_rs_con=seach&_rs_act=keyword_search', headers=headers)
  res.raise_for_status()

  soup = BeautifulSoup(res.text, "lxml")
  companyCards = soup.select(".result_card")
  if len(companyCards) == 0:
    ws.cell(row=row_index, column=2, value="0.0")
  elif len(companyCards) == 1:
    ratings = soup.select(".result_card .rate_ty02")
    rating = ratings[0].get_text().strip()
    ws.cell(row=row_index, column=2, value=rating)
  else:
    ws.cell(row=row_index, column=2, value="0.0")
    ws.cell(row=row_index, column=3, value=f"{len(companyCards)}개")
  row_index = row_index + 1
wb.save("./result/company2.2(2023-07-29).xlsx")