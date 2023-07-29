import requests
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime
import re

today = datetime.today().strftime("%Y.%m.%d %H.%M.%S")
URLS = []
PAGE_COUNT = 11 # 검색 결과의 전체 페이지 수를 입력
for index in range(1,PAGE_COUNT + 1,1):
  # 『 URLS.append(f'여기') 』 여기에 url 입력. url에서『 job-category?page={index}& 』 또는 『 recruitPage={index} 』형태로 입력.
  URLS.append(f'https://www.saramin.co.kr/zf_user/search?loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&cat_kewd=236%2C277&company_cd=0%2C1%2C2%2C3%2C4%2C5%2C6%2C7%2C9%2C10&exp_cd=1&exp_none=y&edu_min=6&edu_max=10&edu_none=y&panel_type=&search_optional_item=y&search_done=y&panel_count=y&preview=y&recruitPage={index}&recruitSort=relation&recruitPageCount=40&inner_com_type=&searchword=&show_applied=&quick_apply=&except_read=&ai_head_hunting=&mainSearch=n')

if os.path.isfile("./result/company2.2(2023-07-29).xlsx") == False:
  wb = Workbook()
else:
  wb = load_workbook("./result/company2.2(2023-07-29).xlsx")
ws = wb.active
ws.title = "기업 별점 시트(" + str(today) + ")"
ws["A1"] = "기업"
ws["B1"] = "별점"
ws["C1"] = "개수"
ws["D1"] = "공고명"
ws["E1"] = "채용 기한"

row_index_A1 = 2
row_index_D1 = 2
row_index_E1 = 2
headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
keywords = ["(주)", "㈜"]
for url in URLS:
    res = requests.get(url, headers=headers)
    res.raise_for_status()

    soup = BeautifulSoup(res.text, "lxml")
    companyNames = soup.select(".company_nm .str_tit")
    companyTitles = soup.select(".area_job .job_tit")
    companyDates = soup.select(".job_date .date")
    if not companyNames:
      companyNames = soup.select(".area_corp .corp_name")

    for companyName in companyNames:
        value = companyName.get_text().strip()
        if any(keyword in value for keyword in keywords):
          value = re.sub("(\(주\)|㈜)", "", value)
        ws.cell(row=row_index_A1, column=1, value=value)
        row_index_A1 = row_index_A1 + 1
    for companyTitle in companyTitles:
        value = companyTitle.get_text().strip()
        ws.cell(row=row_index_D1, column=4, value=value)
        row_index_D1 = row_index_D1 + 1
    for companyDate in companyDates:
        value = companyDate.get_text().strip()
        ws.cell(row=row_index_E1, column=5, value=value)
        row_index_E1 = row_index_E1 + 1
wb.save("./result/company2.2(2023-07-29).xlsx")