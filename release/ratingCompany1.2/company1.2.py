# pip install beautifulsoup4 : 스크래핑을 위한 패키지
# pip install lxml : 구문 분석 파서하는 패키지
# pip install openpyxl : 엑셀 사용하는 패키지
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import os.path
from datetime import datetime

today = datetime.today().strftime("%Y.%m.%d %H.%M.%S")
URLS = ["https://www.saramin.co.kr/zf_user/jobs/list/job-category?cat_kewd=277%2C279%2C236&loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&exp_cd=1&exp_none=y&edu_none=y&edu_min=6&edu_max=10&panel_type=&search_optional_item=y&search_done=y&panel_count=y&preview=y&page=1&page_count=100&sort=RD",
        "https://www.saramin.co.kr/zf_user/jobs/list/job-category?page=2&cat_kewd=277%2C279%2C236&loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&exp_cd=1&exp_none=y&edu_none=y&edu_min=6&edu_max=10&search_optional_item=y&search_done=y&panel_count=y&preview=y&page_count=100&sort=RD&isAjaxRequest=1&type=job-category&is_param=1&isSearchResultEmpty=1&isSectionHome=0&searchParamCount=8#searchTitle",
        "https://www.saramin.co.kr/zf_user/jobs/list/job-category?page=3&cat_kewd=277%2C279%2C236&loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&exp_cd=1&exp_none=y&edu_none=y&edu_min=6&edu_max=10&search_optional_item=y&search_done=y&panel_count=y&preview=y&page_count=100&sort=RD&isAjaxRequest=0&type=job-category&is_param=1&isSearchResultEmpty=1&isSectionHome=0&searchParamCount=8&tab=job-category#searchTitle",
        "https://www.saramin.co.kr/zf_user/jobs/list/job-category?page=4&cat_kewd=277%2C279%2C236&loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&exp_cd=1&exp_none=y&edu_none=y&edu_min=6&edu_max=10&search_optional_item=y&search_done=y&panel_count=y&preview=y&page_count=100&sort=RD&isAjaxRequest=0&type=job-category&is_param=1&isSearchResultEmpty=1&isSectionHome=0&searchParamCount=8&tab=job-category#searchTitle",
        "https://www.saramin.co.kr/zf_user/jobs/list/job-category?page=5&cat_kewd=277%2C279%2C236&loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102170%2C102140&exp_cd=1&exp_none=y&edu_none=y&edu_min=6&edu_max=10&search_optional_item=y&search_done=y&panel_count=y&preview=y&page_count=100&sort=RD&isAjaxRequest=0&type=job-category&is_param=1&isSearchResultEmpty=1&isSectionHome=0&searchParamCount=8&tab=job-category#searchTitle"];
if os.path.isfile("./result/company1.2.xlsx") == False:
  wb = Workbook()
else:
  wb = load_workbook("./result/company1.2.xlsx")
ws = wb.create_sheet()
ws.title = "기업 별점 시트(" + str(today) + ")"
ws["A1"] = "기업"
ws["B1"] = "별점"

row_index = 2
for url in URLS:
    headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
    res = requests.get(url, headers=headers)
    res.raise_for_status()

    soup = BeautifulSoup(res.text, "lxml")
    companys = soup.select(".company_nm .str_tit");

    for company in companys:
        value = company.get_text().strip()
        ws.cell(row=row_index, column=1, value=value)
        row_index = row_index + 1
wb.save("./result/company1.2.xlsx")