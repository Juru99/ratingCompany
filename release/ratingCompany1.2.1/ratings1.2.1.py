import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

options = ChromeOptions()
user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
options.add_argument('user-agent=' + user_agent)
options.add_experimental_option("detach", True) # 브라우저 꺼짐 방지 코드

# 크롬 드라이버 최신 버전 설정
service = ChromeService(executable_path=ChromeDriverManager().install())
# chrome driver
browser = webdriver.Chrome(service=service, options=options)
time.sleep(2) #위 페이지가 모두 열릴 때까지 2초 대기

wb = load_workbook('./result/2023-07-26.xlsx')
ws = wb.active
rowIndex = 2
keywords = []
while not ws.cell(row = rowIndex, column = 1).value is None:
  keywords.append(ws.cell(row = rowIndex, column = 1).value)
  rowIndex = rowIndex + 1
rowIndex = 2
for keyword in keywords:
  browser.get("https://www.jobplanet.co.kr/search?query="+ keyword + "&category=search_new&search_keyword_hint_id=&_rs_con=seach&_rs_act=keyword_search")
  try:
    ratings = browser.find_element(By.CLASS_NAME, 'rate_ty02')
    name = browser.find_element(By.XPATH, '//*[@id="mainContents"]/div[1]/div/div[2]/div[1]/div[1]/a')
    if keyword.replace('(주)','') == name.text.replace('(주)',''):
      ws.cell(row = rowIndex, column = 2, value=ratings.text)
    else:
      ws.cell(row = rowIndex, column = 2, value="0.0")
  except:
    ws.cell(row = rowIndex, column = 2, value="0.0")
  rowIndex = rowIndex + 1
wb.save("./result/2023-07-26.xlsx")