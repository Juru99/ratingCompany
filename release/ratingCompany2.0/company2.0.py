import requests
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime

today = datetime.today().strftime("%Y.%m.%d %H.%M.%S")
URLS = []
PAGE_COUNT = 8 # 100개 단위 검색 시 나오는 총 페이지 개수 작성
for index in range(1,PAGE_COUNT + 1,1):
  # 『 URLS.append(f'여기') 』 여기에 url 입력. 『 job-category?page=요기& 』 요기에 {index} 입력
  URLS.append(f'https://www.saramin.co.kr/zf_user/jobs/list/job-category?page={index}&cat_kewd=236%2C277%2C235&loc_mcd=101000%2C108000&loc_cd=102150%2C102160%2C102140%2C102170&exp_cd=1&exp_none=y&edu_min=6&edu_max=10&edu_none=y&search_optional_item=y&search_done=y&panel_count=y&preview=y&sort=RD&page_count=100&isAjaxRequest=0&type=job-category&is_param=1&isSearchResultEmpty=1&isSectionHome=0&searchParamCount=8&tab=job-category#searchTitle')

if os.path.isfile("./result/company2.0(2023-07-29).xlsx") == False:
  wb = Workbook()
else:
  wb = load_workbook("./result/company2.0(2023-07-29).xlsx")
ws = wb.active
ws.title = "기업 별점 시트(" + str(today) + ")"
ws["A1"] = "기업"
ws["B1"] = "별점"
ws["C1"] = "개수"

row_index = 2
headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"}
for url in URLS:
    res = requests.get(url, headers=headers)
    res.raise_for_status()

    soup = BeautifulSoup(res.text, "lxml")
    companys = soup.select(".company_nm .str_tit")

    for company in companys:
        value = company.get_text().strip()
        ws.cell(row=row_index, column=1, value=value)
        row_index = row_index + 1
wb.save("./result/company2.0(2023-07-29).xlsx")